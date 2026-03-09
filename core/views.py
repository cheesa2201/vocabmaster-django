# BUG FIX: Bỏ "from urllib import request" — nó shadow Django request parameter
# làm toàn bộ view crash với TypeError
import json
import random

from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import load_workbook

from .forms import ExcelImportForm, FileUploadForm, RegisterForm, WordForm
from .models import (LANGUAGE_CHOICES, MAX_WORDS_PER_USER, FileUpload,
                     QuizAnswer, QuizSession, Word)


# ── Auth ──────────────────────────────────────────────────────────────────────

def register(request):
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, "Đăng ký thành công! Chào mừng bạn.")
            return redirect('dashboard')
    else:
        form = RegisterForm()
    return render(request, 'register.html', {'form': form})


# ── Dashboard ─────────────────────────────────────────────────────────────────

@login_required
def dashboard(request):
    total = Word.objects.filter(user=request.user).count()
    stats = {
        code: {
            'label': label,
            'total': Word.objects.filter(user=request.user, lang=code).count(),
            'mastered': Word.objects.filter(user=request.user, lang=code, mastery=2).count(),
        }
        for code, label in LANGUAGE_CHOICES
    }
    recent_files = FileUpload.objects.filter(
        user=request.user,
        expires_at__gt=timezone.now()
    ).order_by('-uploaded_at')[:3]

    from django.db.models import Count
    mastery_qs = (
        Word.objects.filter(user=request.user)
        .values('mastery')
        .annotate(cnt=Count('id'))
    )
    mastery_counts = {0: 0, 1: 0, 2: 0}
    for row in mastery_qs:
        mastery_counts[row['mastery']] = row['cnt']

    return render(request, 'dashboard.html', {
        'total': total,
        'stats': stats,
        'max_words': MAX_WORDS_PER_USER,
        'recent_files': recent_files,
        'mastery_counts': mastery_counts,
    })


# ── Import (xlsx + docx) ──────────────────────────────────────────────────────

@login_required
def import_excel(request):
    if request.method == 'POST':
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            ext = file.name.rsplit('.', 1)[-1].lower()

            current_count = Word.objects.filter(user=request.user).count()
            if current_count >= MAX_WORDS_PER_USER:
                messages.error(request, f"Đã đạt giới hạn {MAX_WORDS_PER_USER:,} từ!")
                return redirect('word_list')

            slots_remaining = MAX_WORDS_PER_USER - current_count
            valid_langs = [code for code, _ in LANGUAGE_CHOICES]
            count_created = count_updated = count_skipped = 0

            def _upsert(lang, word, reading, meaning, wtype, example):
                nonlocal count_created, count_updated, count_skipped
                lang = lang.strip().lower()
                word = word.strip()
                meaning = meaning.strip()
                if lang not in valid_langs or not word or not meaning:
                    count_skipped += 1
                    return False
                existing = Word.objects.filter(
                    user=request.user, lang=lang, word=word
                ).first()
                if existing:
                    existing.reading = reading.strip()
                    existing.meaning = meaning
                    existing.word_type = wtype.strip()
                    existing.example = example.strip()
                    existing.save(update_fields=['reading', 'meaning', 'word_type', 'example'])
                    count_updated += 1
                else:
                    Word.objects.create(
                        user=request.user, lang=lang, word=word,
                        reading=reading.strip(), meaning=meaning,
                        word_type=wtype.strip(), example=example.strip(),
                        mastery=0, last_review=timezone.now(),
                    )
                    count_created += 1
                return True

            if ext == 'xlsx':
                wb = load_workbook(file, read_only=True)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if count_created >= slots_remaining:
                        messages.warning(request, f"Dừng sớm: đạt giới hạn {MAX_WORDS_PER_USER:,} từ.")
                        break
                    if not row or len(row) < 4:
                        count_skipped += 1
                        continue
                    _upsert(
                        str(row[0] or ''), str(row[1] or ''), str(row[2] or ''),
                        str(row[3] or ''),
                        str(row[4]) if len(row) > 4 and row[4] else '',
                        str(row[5]) if len(row) > 5 and row[5] else '',
                    )
                wb.close()

            elif ext == 'docx':
                try:
                    from mammoth import extract_raw_text  # type: ignore
                    text = extract_raw_text(file).value
                except ImportError:
                    messages.error(request, "Thiếu thư viện mammoth. Chạy: pip install mammoth")
                    return redirect('import_excel')

                for line in text.split('\n'):
                    if count_created >= slots_remaining:
                        break
                    parts = line.split('|')
                    if len(parts) < 4:
                        count_skipped += 1
                        continue
                    _upsert(
                        parts[0], parts[1], parts[2], parts[3],
                        parts[4] if len(parts) > 4 else '',
                        parts[5] if len(parts) > 5 else '',
                    )
            else:
                messages.error(request, "Chỉ hỗ trợ .xlsx hoặc .docx")
                return redirect('import_excel')

            messages.success(
                request,
                f"Import xong: +{count_created} mới, {count_updated} cập nhật, {count_skipped} bỏ qua."
            )
            return redirect('word_list')
    else:
        form = ExcelImportForm()

    current_count = Word.objects.filter(user=request.user).count()
    return render(request, 'import_excel.html', {
        'form': form,
        'current_count': current_count,
        'max_words': MAX_WORDS_PER_USER,
        'slots_remaining': MAX_WORDS_PER_USER - current_count,
    })


# ── Word List ─────────────────────────────────────────────────────────────────

@login_required
def word_list(request):
    lang_filter = request.GET.get('lang', '')
    mastery_filter = request.GET.get('mastery', '')

    words = Word.objects.filter(user=request.user)
    if lang_filter in [c for c, _ in LANGUAGE_CHOICES]:
        words = words.filter(lang=lang_filter)
    if mastery_filter in ['0', '1', '2']:
        words = words.filter(mastery=int(mastery_filter))

    return render(request, 'word_list.html', {
        'words': words,
        'lang_choices': LANGUAGE_CHOICES,
        'lang_filter': lang_filter,
        'mastery_filter': mastery_filter,
        'total': words.count(),
    })


# ── Flashcard ─────────────────────────────────────────────────────────────────

@login_required
def flashcard(request):
    lang_filter = request.GET.get('lang', '')
    qs = Word.objects.filter(user=request.user)
    if lang_filter in [c for c, _ in LANGUAGE_CHOICES]:
        qs = qs.filter(lang=lang_filter)
    words = list(qs.order_by('?')[:15])
    return render(request, 'flashcard.html', {
        'words': words,
        'lang_choices': LANGUAGE_CHOICES,
        'lang_filter': lang_filter,
    })


# ── Update mastery ────────────────────────────────────────────────────────────

@login_required
def update_mastery(request, word_id):
    if request.method == 'POST':
        word = Word.objects.filter(id=word_id, user=request.user).first()
        if word:
            new_mastery = int(request.POST.get('mastery', 0))
            if new_mastery in [0, 1, 2]:
                word.mastery = new_mastery
                word.last_review = timezone.now()
                word.save(update_fields=['mastery', 'last_review'])
    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))


# ── Quiz ──────────────────────────────────────────────────────────────────────

@login_required
def quiz_home(request):
    word_counts = {
        code: Word.objects.filter(user=request.user, lang=code).count()
        for code, _ in LANGUAGE_CHOICES
    }
    total = sum(word_counts.values())
    return render(request, 'quiz_home.html', {
        'lang_choices': LANGUAGE_CHOICES,
        'word_counts': word_counts,
        'word_counts_js': json.dumps(word_counts),
        'total': total,
    })


@login_required
def quiz_start(request):
    if request.method != 'POST':
        return redirect('quiz_home')

    quiz_type = request.POST.get('quiz_type', 'fill')
    lang = request.POST.get('lang', '')
    num_q = max(5, min(int(request.POST.get('num_questions', 10)), 20))

    qs = Word.objects.filter(user=request.user)
    if lang:
        qs = qs.filter(lang=lang)

    if qs.count() < 4:
        messages.error(request, "Cần ít nhất 4 từ vựng để bắt đầu quiz!")
        return redirect('quiz_home')

    word_ids = list(qs.order_by('?').values_list('id', flat=True)[:num_q])
    session = QuizSession.objects.create(
        user=request.user, quiz_type=quiz_type,
        lang=lang, total_questions=len(word_ids),
    )
    request.session[f'quiz_{session.id}_words'] = word_ids
    return redirect('quiz_question', session_id=session.id, q_index=0)


@login_required
def quiz_question(request, session_id, q_index):
    try:
        session = QuizSession.objects.get(id=session_id, user=request.user)
    except QuizSession.DoesNotExist:
        messages.error(request, "Không tìm thấy phiên quiz.")
        return redirect('quiz_home')

    word_ids = request.session.get(f'quiz_{session.id}_words', [])
    if not word_ids:
        messages.error(request, "Phiên quiz hết hạn. Hãy bắt đầu lại.")
        return redirect('quiz_home')

    if q_index >= len(word_ids):
        return redirect('quiz_result', session_id=session.id)

    word = Word.objects.get(id=word_ids[q_index])

    if request.method == 'POST':
        user_answer = request.POST.get('answer', '').strip()
        is_correct = _check_answer(session.quiz_type, user_answer, word)

        QuizAnswer.objects.create(
            session=session, word=word,
            user_answer=user_answer, is_correct=is_correct,
        )
        if is_correct:
            session.correct_count += 1
            session.save(update_fields=['correct_count'])

        _update_mastery_quiz(word, is_correct)

        request.session[f'quiz_{session.id}_last'] = {
            'is_correct': is_correct,
            'correct_answer': word.meaning,
            'user_answer': user_answer,
            'word': word.word,
            'reading': word.reading,
        }
        return redirect('quiz_question', session_id=session.id, q_index=q_index + 1)

    last_result = request.session.pop(f'quiz_{session.id}_last', None)
    context = {
        'session': session,
        'word': word,
        'q_index': q_index,
        'q_number': q_index + 1,
        'total': len(word_ids),
        'progress_percent': round(q_index / len(word_ids) * 100),
        'last_result': last_result,
    }
    if session.quiz_type == 'multiple':
        context['choices'] = _get_choices(word, request.user, word_ids)

    return render(request, 'quiz_question.html', context)


@login_required
def quiz_result(request, session_id):
    try:
        session = QuizSession.objects.get(id=session_id, user=request.user)
    except QuizSession.DoesNotExist:
        return redirect('quiz_home')

    if not session.finished_at:
        session.finished_at = timezone.now()
        session.save(update_fields=['finished_at'])

    request.session.pop(f'quiz_{session.id}_words', None)
    answers = session.answers.select_related('word').all()

    return render(request, 'quiz_result.html', {
        'session': session,
        'answers': answers,
        'score_percent': session.score_percent,
        'duration': session.duration_seconds,
    })


@login_required
def quiz_history(request):
    sessions = QuizSession.objects.filter(
        user=request.user,
        finished_at__isnull=False,
    )[:30]
    return render(request, 'quiz_history.html', {'sessions': sessions})


# ── File upload ───────────────────────────────────────────────────────────────

@login_required
def upload_file(request):
    if request.method == 'POST':
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            f = form.save(commit=False)
            f.user = request.user
            f.save()
            messages.success(request, "Upload thành công!")
            return redirect('file_list')
    else:
        form = FileUploadForm()
    return render(request, 'upload_file.html', {'form': form})


@login_required
def file_list(request):
    files = FileUpload.objects.filter(
        user=request.user,
        expires_at__gt=timezone.now()
    ).order_by('-uploaded_at')
    return render(request, 'file_list.html', {'files': files})


@login_required
def view_file(request, pk):
    file = get_object_or_404(FileUpload, pk=pk, user=request.user)

    # BUG FIX: Check file hết hạn
    if file.is_expired:
        messages.error(request, "File này đã hết hạn và không thể xem.")
        return redirect('file_list')

    return render(request, 'view_file.html', {'file': file})


@login_required
def save_note(request, pk):
    """Lưu ghi chú cho file — POST only."""
    if request.method != 'POST':
        return redirect('file_list')

    file = get_object_or_404(FileUpload, pk=pk, user=request.user)
    if file.is_expired:
        return JsonResponse({'ok': False, 'error': 'File đã hết hạn'}, status=403)

    note_text = request.POST.get('note', '').strip()
    # Lưu đơn giản: [{text, saved_at}]
    notes = file.notes_json or []
    # BUG FIX: dùng append thay vì gán mới → không mất ghi chú cũ
    notes.append({'text': note_text, 'saved_at': timezone.now().isoformat()})
    file.notes_json = notes
    file.save(update_fields=['notes_json'])

    messages.success(request, "Đã lưu ghi chú!")
    return redirect('view_file', pk=pk)


# ── Quiz helpers ──────────────────────────────────────────────────────────────

def _check_answer(quiz_type, user_answer, word):
    if not user_answer:
        return False
    if quiz_type == 'fill':
        correct_parts = [p.strip().lower() for p in word.meaning.replace('/', ',').split(',')]
        given_parts = [p.strip().lower() for p in user_answer.replace('/', ',').split(',')]
        return any(g for g in given_parts if any(g in c or c in g for c in correct_parts))
    elif quiz_type == 'multiple':
        return user_answer.strip() == word.meaning.strip()
    return False


def _get_choices(word, user, exclude_ids):
    distractors = list(
        Word.objects.filter(user=user, lang=word.lang)
        .exclude(id__in=exclude_ids).exclude(id=word.id)
        .order_by('?').values_list('meaning', flat=True)[:3]
    )
    if len(distractors) < 3:
        extra = list(
            Word.objects.filter(user=user)
            .exclude(id__in=exclude_ids).exclude(id=word.id)
            .order_by('?').values_list('meaning', flat=True)[:3 - len(distractors)]
        )
        distractors += extra
    choices = distractors + [word.meaning]
    random.shuffle(choices)
    return choices


def _update_mastery_quiz(word, correct):
    word.mastery = min(word.mastery + 1, 2) if correct else max(word.mastery - 1, 0)
    word.last_review = timezone.now()
    word.save(update_fields=['mastery', 'last_review'])


# ── Health check (Railway) ────────────────────────────────────────────────────

from django.http import HttpResponse

def health_check(request):
    """Public endpoint cho Railway healthcheck — không cần login."""
    return HttpResponse("ok", content_type="text/plain", status=200)